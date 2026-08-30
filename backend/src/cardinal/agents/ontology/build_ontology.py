"""Convert the Excel ontology workbook into billing.ttl and refresh the HTML map.

Excel is the single source of truth. Run this after editing the workbook:

    python build_ontology.py

It checks referential integrity (every dependency and obligation points at a real
field), then regenerates billing.ttl (which imports core.ttl) and billing-ontology.html.
Needs only openpyxl. Loading/reasoning over the .ttl (rdflib/owlrl/pyshacl) is a
separate runtime concern; this converter does not need them.
"""
from __future__ import annotations
import os
import sys
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "billing_ontology.xlsx")
TTL = os.path.join(HERE, "billing.ttl")
HTML = os.path.join(HERE, "billing-ontology.html")

PRIORITY = {"regulatory": "card:Regulatory", "structural": "card:Structural", "refinement": "card:Refinement"}


# ----------------------------- read -----------------------------

def read_sheet(ws) -> list[dict]:
    heads = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value in (None, ""):
            continue
        rows.append({heads[c - 1]: ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)})
    return rows


def load():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    return {
        "grains": read_sheet(wb["Grains"]),
        "fields": read_sheet(wb["Fields"]),
        "deps": read_sheet(wb["Dependencies"]),
        "invs": read_sheet(wb["Invariants"]),
        "obls": read_sheet(wb["Obligations"]),
    }


def g(row, key):
    v = row.get(key)
    return "" if v is None else str(v).strip()


# ----------------------------- validate -----------------------------

def validate(data) -> list[str]:
    errs = []
    field_ids = {g(f, "id") for f in data["fields"]}
    grain_ids = {g(gr, "id") for gr in data["grains"]}
    inv_ids = {g(i, "id") for i in data["invs"]}

    for f in data["fields"]:
        if g(f, "grain") not in grain_ids:
            errs.append(f"field {g(f,'id')}: unknown grain '{g(f,'grain')}'")
    for d in data["deps"]:
        s, t = g(d, "source (used)"), g(d, "target (calculated from source)")
        if s not in field_ids:
            errs.append(f"dependency {g(d,'id')}: unknown source '{s}'")
        if t not in field_ids:
            errs.append(f"dependency {g(d,'id')}: unknown target '{t}'")
        if g(d, "lag") not in ("0", "1", "2", "3", "4", "5", "6"):
            errs.append(f"dependency {g(d,'id')}: lag '{g(d,'lag')}' is not a small integer")
    for o in data["obls"]:
        if g(o, "triggered_by") and g(o, "triggered_by") not in field_ids:
            errs.append(f"obligation {g(o,'id')}: unknown triggered_by '{g(o,'triggered_by')}'")
        for fid in [x.strip() for x in g(o, "requires_fields").split(",") if x.strip()]:
            if fid not in field_ids:
                errs.append(f"obligation {g(o,'id')}: requires unknown field '{fid}'")
        for iid in [x.strip() for x in g(o, "requires_invariants").split(",") if x.strip()]:
            if iid not in inv_ids:
                errs.append(f"obligation {g(o,'id')}: requires unknown invariant '{iid}'")
    return errs


# ----------------------------- emit ttl -----------------------------

def lit(s: str) -> str:
    return '"' + s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n") + '"'


def emit_ttl(data) -> str:
    L = []
    L.append("# GENERATED FILE - do not edit by hand.")
    L.append("# Source of truth: billing_ontology.xlsx. Regenerate with build_ontology.py.")
    L.append("")
    L.append("@prefix card: <https://cardinal.dev/onto#> .")
    L.append("@prefix bill: <https://cardinal.dev/onto/billing#> .")
    L.append("@prefix owl:  <http://www.w3.org/2002/07/owl#> .")
    L.append("@prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .")
    L.append("@prefix skos: <http://www.w3.org/2004/02/skos/core#> .")
    L.append("@prefix dct:  <http://purl.org/dc/terms/> .")
    L.append("")
    L.append("<https://cardinal.dev/onto/billing> a owl:Ontology ;")
    L.append("    owl:imports <https://cardinal.dev/onto> ;")
    L.append('    dct:title "Cardinal billing ontology (generated)"@en .')
    L.append("")

    L.append("# ---- Grains ----")
    for gr in data["grains"]:
        L.append(f'bill:{g(gr,"id")} a card:Grain ; skos:prefLabel {lit(g(gr,"label"))}@en ; card:grainKey {lit(g(gr,"key"))} .')
    L.append("")

    L.append("# ---- Fields ----")
    for f in data["fields"]:
        fid = g(f, "id")
        cls = "card:DerivedField" if g(f, "kind") == "derived" else "card:Field"
        parts = [f"bill:{fid} a {cls} ;"]
        parts.append(f'    skos:prefLabel {lit(g(f,"label"))}@en ;')
        parts.append(f'    card:onGrain bill:{g(f,"grain")} ;')
        if g(f, "dtype"):
            parts.append(f'    card:dtype {lit(g(f,"dtype"))} ;')
        if g(f, "kind") == "derived":
            parts.append("    card:derived true ;")
            if g(f, "formula"):
                parts.append(f'    card:formula {lit(g(f,"formula"))} ;')
        if g(f, "distribution"):
            parts.append(f'    card:distFamily card:{g(f,"distribution")} ;')
        if g(f, "gate"):
            parts.append(f'    card:gate {lit(g(f,"gate"))} ;')
        if g(f, "priority") in PRIORITY:
            parts.append(f'    card:priority {PRIORITY[g(f,"priority")]} ;')
        if g(f, "source"):
            parts.append(f'    dct:source {lit(g(f,"source"))}@en ;')
        if g(f, "why"):
            parts.append(f'    card:why {lit(g(f,"why"))} ;')
        if g(f, "ask"):
            parts.append(f'    card:ask {lit(g(f,"ask"))} ;')
        parts[-1] = parts[-1][:-2] + " ."  # replace trailing ' ;' with ' .'
        L.append("\n".join(parts))
    L.append("")

    L.append("# ---- Dependencies (+ directlyDependsOn edges for the reasoner) ----")
    for d in data["deps"]:
        did = g(d, "id")
        s = g(d, "source (used)")
        t = g(d, "target (calculated from source)")
        parts = [f"bill:{did} a card:Dependency ;"]
        parts.append(f"    card:source bill:{s} ; card:target bill:{t} ;")
        parts.append(f'    card:lag {g(d,"lag") or "0"} ;')
        if g(d, "mechanism"):
            parts.append(f'    card:mechanism card:{g(d,"mechanism")} ;')
        if g(d, "priority") in PRIORITY:
            parts.append(f'    card:priority {PRIORITY[g(d,"priority")]} ;')
        if g(d, "source (cite)"):
            parts.append(f'    dct:source {lit(g(d,"source (cite)"))}@en ;')
        if g(d, "why"):
            parts.append(f'    card:why {lit(g(d,"why"))} ;')
        parts[-1] = parts[-1][:-2] + " ."
        L.append("\n".join(parts))
        L.append(f"bill:{t} card:directlyDependsOn bill:{s} .")
    L.append("")

    L.append("# ---- Invariants ----")
    for i in data["invs"]:
        iid = g(i, "id")
        parts = [f"bill:{iid} a card:Invariant ;"]
        parts.append(f'    skos:prefLabel {lit(g(i,"label"))}@en ;')
        parts.append(f'    card:expr {lit(g(i,"rule"))} ;')
        if g(i, "severity"):
            parts.append(f'    card:severity {lit(g(i,"severity"))} ;')
        if g(i, "grain"):
            parts.append(f'    card:appliesToGrain bill:{g(i,"grain")} ;')
        if g(i, "priority") in PRIORITY:
            parts.append(f'    card:priority {PRIORITY[g(i,"priority")]} ;')
        if g(i, "source"):
            parts.append(f'    dct:source {lit(g(i,"source"))}@en ;')
        parts[-1] = parts[-1][:-2] + " ."
        L.append("\n".join(parts))
    L.append("")

    L.append("# ---- Obligations ----")
    for o in data["obls"]:
        oid = g(o, "id")
        cls = "card:RegulatoryObligation" if g(o, "priority") == "regulatory" else "card:Obligation"
        parts = [f"bill:{oid} a {cls} ;"]
        parts.append(f'    skos:prefLabel {lit(g(o,"label"))}@en ;')
        if g(o, "triggered_by"):
            parts.append(f'    card:triggeredByField bill:{g(o,"triggered_by")} ;')
        reqf = [x.strip() for x in g(o, "requires_fields").split(",") if x.strip()]
        if reqf:
            parts.append("    card:requiresField " + " , ".join(f"bill:{x}" for x in reqf) + " ;")
        reqi = [x.strip() for x in g(o, "requires_invariants").split(",") if x.strip()]
        if reqi:
            parts.append("    card:requiresInvariant " + " , ".join(f"bill:{x}" for x in reqi) + " ;")
        if g(o, "priority") in PRIORITY:
            parts.append(f'    card:priority {PRIORITY[g(o,"priority")]} ;')
        if g(o, "source"):
            parts.append(f'    dct:source {lit(g(o,"source"))}@en ;')
        if g(o, "ask"):
            parts.append(f'    card:ask {lit(g(o,"ask"))} ;')
        if g(o, "why"):
            parts.append(f'    card:why {lit(g(o,"why"))} ;')
        parts[-1] = parts[-1][:-2] + " ."
        L.append("\n".join(parts))
    L.append("")
    return "\n".join(L)


# ----------------------------- emit html map -----------------------------

def xml_escape(s: str) -> str:
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
             .replace('"', "&quot;"))


def layer_nodes(fields, deps):
    ids = [g(f, "id") for f in fields]
    lag0 = [(g(d, "source (used)"), g(d, "target (calculated from source)"))
            for d in deps if (g(d, "lag") or "0") == "0"]
    indeg = {i: 0 for i in ids}
    out = {i: [] for i in ids}
    for s, t in lag0:
        if s in indeg and t in indeg:
            indeg[t] += 1
            out[s].append(t)
    layer = {i: 0 for i in ids}
    preds = {i: [] for i in ids}
    queue = [i for i in ids if indeg[i] == 0]
    seen = set(queue)
    while queue:
        n = queue.pop(0)
        for t in out[n]:
            preds[t].append(n)
            layer[t] = max(layer[t], layer[n] + 1)
            indeg[t] -= 1
            if indeg[t] == 0 and t not in seen:
                queue.append(t); seen.add(t)
    # pull pure-input nodes rightward, next to their earliest consumer,
    # so inputs don't all pile into column 0 with long crossing edges
    for i in ids:
        if not preds[i] and out[i]:
            layer[i] = max(0, min(layer[t] for t in out[i]) - 1)
    return layer


def emit_html(data) -> str:
    fields = data["fields"]
    deps = data["deps"]
    by_id = {g(f, "id"): f for f in fields}
    layer = layer_nodes(fields, deps)
    maxlayer = max(layer.values()) if layer else 0

    # positions
    cols = {}
    for i in sorted(layer, key=lambda k: (layer[k], k)):
        cols.setdefault(layer[i], []).append(i)
    W, H = 210, 56
    XSP, YSP = 250, 88
    pos = {}
    maxrows = 0
    for lyr, nodes in cols.items():
        maxrows = max(maxrows, len(nodes))
        for idx, nid in enumerate(nodes):
            pos[nid] = (40 + lyr * XSP, 50 + idx * YSP)
    canvas_w = 40 + maxlayer * XSP + W + 40
    canvas_h = 60 + maxrows * YSP + 20

    def color(fid):
        f = by_id[fid]
        if g(f, "priority") == "regulatory":
            return "#fdeccd", "#e0a94a"
        if g(f, "kind") == "derived":
            return "#d8eafa", "#82add6"
        return "#f3f5f7", "#b6c0ca"

    svg = [f'<svg viewBox="0 0 {canvas_w} {canvas_h}" xmlns="http://www.w3.org/2000/svg" '
           f'role="img" aria-label="Auto-generated map of billing fields and dependencies">']
    svg.append('<defs>')
    for mid, col in (("head", "#3a4757"), ("lagHead", "#7a56b0"), ("gateHead", "#c98a17")):
        svg.append(f'<marker id="{mid}" markerWidth="10" markerHeight="10" refX="8" refY="3" '
                   f'orient="auto" markerUnits="strokeWidth"><path d="M0 0 L0 6 L9 3 Z" fill="{col}"/></marker>')
    svg.append('</defs>')

    # edges
    for d in deps:
        s = g(d, "source (used)"); t = g(d, "target (calculated from source)")
        if s not in pos or t not in pos:
            continue
        sx, sy = pos[s]; tx, ty = pos[t]
        scy = sy + H / 2; tcy = ty + H / 2
        x1 = sx + W; x2 = tx
        lag = g(d, "lag") or "0"
        mech = g(d, "mechanism")
        if lag != "0":
            cls, mk, lab = "#7a56b0", "lagHead", f"waits {lag} mo"
        elif mech == "Gate":
            cls, mk, lab = "#c98a17", "gateHead", ""
        else:
            cls, mk, lab = "#3a4757", "head", ""
        dash = ' stroke-dasharray="6 5"' if mech == "Gate" else ""
        if layer.get(t, 0) > layer.get(s, 0):
            path = f"M{x1} {scy:.0f} C{x1+40} {scy:.0f}, {x2-40} {tcy:.0f}, {x2} {tcy:.0f}"
        else:  # back or same layer: dip below
            path = f"M{x1} {scy:.0f} C{x1+40} {scy+110:.0f}, {x2-40} {tcy+110:.0f}, {x2} {tcy:.0f}"
        svg.append(f'<path d="{path}" fill="none" stroke="{cls}" stroke-width="1.8"{dash} marker-end="url(#{mk})"/>')
        if lab:
            mx, my = x1 + (x2 - x1) * 0.28, scy + (tcy - scy) * 0.28 - 6
            svg.append(f'<text x="{mx:.0f}" y="{my:.0f}" font-size="10.5" font-weight="bold" '
                       f'fill="{cls}" text-anchor="middle">{xml_escape(lab)}</text>')

    # nodes
    for fid, (x, y) in pos.items():
        fill, stroke = color(fid)
        f = by_id[fid]
        title = xml_escape(g(f, "label") or fid)
        sub = "you enter" if g(f, "kind") == "entered" else "calculated"
        svg.append(f'<rect x="{x}" y="{y}" width="{W}" height="{H}" rx="9" fill="{fill}" stroke="{stroke}" stroke-width="1.5"/>')
        svg.append(f'<text x="{x+W/2:.0f}" y="{y+24}" font-size="13" font-weight="bold" text-anchor="middle">{title}</text>')
        svg.append(f'<text x="{x+W/2:.0f}" y="{y+41}" font-size="10.5" fill="#51617a" text-anchor="middle">{sub}</text>')
    svg.append("</svg>")
    svg_str = "\n".join(svg)

    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Cardinal billing ontology, in plain words</title>
<style>
* {{ box-sizing:border-box; }}
body {{ margin:0; padding:32px; background:#fff; color:#17212f; font-family:Arial, Helvetica, sans-serif; line-height:1.5; }}
main {{ max-width:1200px; margin:auto; }}
h1 {{ margin:0 0 4px; font-size:28px; }} h2 {{ font-size:18px; margin:28px 0 12px; }}
.sub {{ color:#536174; font-size:15px; margin:0 0 4px; }}
.stamp {{ color:#8794a5; font-size:12px; margin:0 0 24px; }}
.cards {{ display:flex; gap:16px; flex-wrap:wrap; margin-bottom:8px; }}
.card {{ flex:1 1 320px; border:1px solid #d9e1ea; border-radius:10px; padding:16px 18px; }}
.card h3 {{ margin:0 0 6px; font-size:15px; }} .card p {{ margin:0; font-size:14px; color:#3a4757; }}
.card.brain {{ background:#f6f1fb; border-color:#d6c2ea; }} .card.recipe {{ background:#eef7ec; border-color:#bcdcb6; }}
.canvas {{ overflow-x:auto; border:1px solid #d9e1ea; border-radius:8px; margin-top:4px; }}
svg {{ display:block; min-width:{canvas_w}px; background:#fff; }}
.legend {{ display:flex; gap:20px; flex-wrap:wrap; font-size:13px; color:#3a4757; margin:10px 2px 0; }}
.legend span {{ display:inline-flex; align-items:center; gap:7px; }}
.sw {{ width:22px; height:14px; border-radius:3px; display:inline-block; border:1.5px solid; }}
.sw.in {{ background:#f3f5f7; border-color:#b6c0ca; }} .sw.calc {{ background:#d8eafa; border-color:#82add6; }}
.sw.law {{ background:#fdeccd; border-color:#e0a94a; }}
.lline {{ width:26px; height:0; border-top:2px solid; display:inline-block; }}
.story {{ background:#fffaf0; border:1px solid #e7cf9c; border-radius:10px; padding:16px 20px; }}
.story ol {{ margin:8px 0 0; padding-left:20px; }} .story li {{ margin:6px 0; font-size:14px; }}
table.gloss {{ border-collapse:collapse; width:100%; font-size:14px; }}
table.gloss td {{ border-top:1px solid #e6ebf0; padding:8px 10px; vertical-align:top; }}
table.gloss td.term {{ width:200px; font-weight:bold; color:#2a3647; }}
</style></head><body><main>
<h1>How credit card billing works, as a map</h1>
<p class="sub">This is the "brain" the app already knows. Each box is one thing on a statement. Each arrow means "this is worked out from that."</p>
<p class="stamp">Generated from billing_ontology.xlsx by build_ontology.py. {len(fields)} fields, {len(deps)} dependencies. Do not edit this page by hand.</p>
<div class="cards">
  <div class="card brain"><h3>The brain (this map)</h3><p>What the app knows about credit cards and the law, before you say anything. You do not write this. It guides the questions the app asks you.</p></div>
  <div class="card recipe"><h3>Your recipe (comes later)</h3><p>The choices you make in the conversation. The app writes them into a file that tells the generator how to build your data.</p></div>
</div>
<h2>The map</h2>
<div class="legend">
  <span><span class="sw in"></span> You enter this</span>
  <span><span class="sw calc"></span> The app works this out</span>
  <span><span class="sw law"></span> The law requires this</span>
  <span><span class="lline" style="border-color:#3a4757"></span> worked out from</span>
  <span><span class="lline" style="border-color:#7a56b0"></span> uses last month</span>
  <span><span class="lline" style="border-color:#c98a17;border-top-style:dashed"></span> a switch (on/off)</span>
</div>
<div class="canvas">
{svg_str}
</div>
<h2>The one story worth understanding: the grace period</h2>
<div class="story">
<p style="margin:0;font-size:14px;">Follow the amber boxes. This is the rule the law forces in, and the one most people model wrong.</p>
<ol>
<li><b>Did they pay last month's bill in full?</b> The app looks back one month (the purple arrows) and answers yes or no.</li>
<li><b>If yes, the grace period is on.</b> Grace is a no-interest window: new spending is not charged interest.</li>
<li><b>So interest gets switched off.</b> That is the dashed amber arrow. During grace, interest for the month is zero.</li>
<li><b>If no, interest is charged</b> as the average balance times the monthly rate.</li>
</ol>
</div>
<h2>The words, in plain terms</h2>
<table class="gloss">
<tr><td class="term">Ontology (the brain)</td><td>A stored map of how credit card billing works and what the law requires.</td></tr>
<tr><td class="term">Node (box)</td><td>One thing on a statement: a number, or a yes/no.</td></tr>
<tr><td class="term">Edge (arrow)</td><td>"This is worked out from that." The arrow points from what is used to what it produces.</td></tr>
<tr><td class="term">Entered vs worked out</td><td>Gray boxes are values we make up to feed the model. Blue and amber boxes are calculated by a fixed formula.</td></tr>
<tr><td class="term">Lag ("uses last month")</td><td>Some things depend on last month, not this month. Shown as purple arrows.</td></tr>
<tr><td class="term">Gate (a switch)</td><td>One thing switches another on or off. Grace switches interest off. Shown dashed.</td></tr>
<tr><td class="term">Obligation</td><td>A rule the law forces you to include. The app asks about it first.</td></tr>
<tr><td class="term">Invariant</td><td>A check the finished data must always pass. If it breaks, the account is remade.</td></tr>
</table>
</main></body></html>
"""


# ----------------------------- main -----------------------------

def main():
    data = load()
    errs = validate(data)
    if errs:
        print("VALIDATION FAILED:")
        for e in errs:
            print("  -", e)
        print("Nothing written.")
        sys.exit(1)
    with open(TTL, "w", encoding="utf-8") as fh:
        fh.write(emit_ttl(data))
    with open(HTML, "w", encoding="utf-8") as fh:
        fh.write(emit_html(data))
    print("OK. Wrote:")
    print("  ", TTL)
    print("  ", HTML)
    print(f"counts: {len(data['fields'])} fields, {len(data['deps'])} deps, "
          f"{len(data['invs'])} invariants, {len(data['obls'])} obligations")


if __name__ == "__main__":
    main()
